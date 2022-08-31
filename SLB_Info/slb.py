import re
from openpyxl import Workbook
from openpyxl import load_workbook


def openfile():
    f = open('nsconfig.txt', 'r')
    if f:
        s = f.read()
        config = s.split('\n')
    else:
        print('File %s not found' % (file))
    f.close()
    return config

def vserver_parse(config):
    vserver=list()
    p=re.compile('add lb vserver')
    for i in config:
        m=p.search(i)
        if m:
            vserver.append(i)
    return vserver

def service_parse(config):
    service=list()
    p=re.compile('add service')
    for i in config:
        m=p.search(i)
        if m:
            service.append(i)
    return service

def bind_parse(config):
    bind=list()
    p=re.compile('bind lb vserver')
    for i in config:
        m=p.search(i)
        if m:
            bind.append(i)
    return bind

def server_parse(config):
    servers=list()
    p=re.compile('add server')
    for i in config:
        m=p.search(i)
        if m:
            servers.append(i)
    return servers

def vserver_create(vinfo, anchor, num):    
    wb = load_workbook('C:/SLB_info/slb_info.xlsx')
    ws = wb['slb']    
    ws['A'+str(anchor)] = str(num)
    ws['B'+str(anchor)] = vinfo[3] 
    ws['C'+str(anchor)] = vinfo[10]
    ws['D'+str(anchor)] = vinfo[5]
    ws['E'+str(anchor)] = vinfo[6]
    ws['F'+str(anchor)] = vinfo[4]
    num += 1         
    wb.save("slb_info.xlsx")
    wb.close()
    return num

def find_service(v, bind):
    service_name=list()
    p=re.compile(v)
    for i in bind:
        m=p.search(i)
        if m:
            i=' '.join(i.split())
            ps=i.split(' ')
            if ps[3] == v:
                service_name.append(ps[4])
    print(service_name)            
    return service_name
    

def service_info(service_name, service, servers, anchor):
    wb = load_workbook('C:/SLB_info/slb_info.xlsx')
    ws = wb['slb']
    service_num = 1
    if service_name: 
        for i in service_name:
            # print(i)
            p=re.compile(i)            
            for s in service:            
                # print(s)
                m=p.search(s)
                # print(m)
                if m:
                    ps=s.split(' ')
                    # ip = re.findall(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}' , ps[3])
                    if i == ps[2]:                
                        ws['G'+str(anchor)] = ps[5]
                        ws['H'+str(anchor)] = str(service_num)+")"
                        ws['I'+str(anchor)] = ps[2]
                        for server in servers:
                            pp=server.split(' ')
                            # print(pp)
                            # print(ps[3])
                            # print(pp[2])
                            if ps[3] == pp[2]:
                                ws['J'+str(anchor)] = pp[3]                              
                        anchor += 1
                        service_num += 1
                        # breakpoint()
    else:
        ws['H'+str(anchor)] = str(service_num)+")"
        ws['I'+str(anchor)] = "No Service"
        anchor += 1
        service_num += 1
    wb.save("slb_info.xlsx")
    wb.close()           
    return anchor


def cell_merge(anchor, anchor_return):
    wb = load_workbook('C:/SLB_info/slb_info.xlsx')
    ws = wb['slb']
    anchor_end = anchor_return - 1 
    for x in [1, 2, 3, 4, 5, 6]:
        ws.merge_cells(start_row = anchor, start_column=x, end_row = anchor_end, end_column=x)
    wb.save("slb_info.xlsx")
    wb.close()

def not_v_service(service, servers, anchor):
    wb = load_workbook('C:/SLB_info/slb_info.xlsx')
    ws = wb['slb']
    true_service=list()
    for cell in range(3, anchor):
        true_service.append(ws.cell(cell, 9).value)
    for i in true_service:
        if i == "No Service":
            true_service.remove('No Service')
    for i in service:
        ps=i.split(' ')
        ip = re.findall(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}' , ps[3])
        if ps[2] not in true_service:
            ws['G'+str(anchor)] = ps[5]
            ws['H'+str(anchor)] = "1)"
            ws['I'+str(anchor)] = ps[2] 
            # ws['J'+str(anchor)] = str(ip[0])
            for server in servers:
                pp=server.split(' ')
                if ps[3] == pp[2]:
                    ws['J'+str(anchor)] = pp[3] 
            anchor += 1
    wb.save("slb_info.xlsx")
    wb.close()
    return anchor

def service_stat(f_anchor):
    wb = load_workbook('C:/SLB_info/slb_info.xlsx')
    ws = wb['slb']
    full_service=list()
    stat_anchor=3
    for cell in range(3, f_anchor):
        full_service.append(ws.cell(cell, 9).value)
    f = open('stat_service.txt', 'r')
    if f:
        s = f.read()
        stat_service = s.split('\n')
    else:
        print('File %s not found' % (file))
    for i in full_service:
        p=re.compile(i)
        # print(i)
        regex = re.compile("^\"")
        patten = regex.search(i)
        if i == "No Service" or patten:
            ws['M'+str(stat_anchor)] = ""
            stat_anchor += 1
        else:
            for s in stat_service:
                m=p.search(s)
                if m:
                    v = s.strip()
                    ps=v.split(' ')
                    # print(ps)                    
                    if i == ps[0]:
                        ws['M'+str(stat_anchor)] = ps[-1]
                        stat_anchor += 1                                             
    del stat_anchor
    wb.save("slb_info.xlsx")
    wb.close()

def main():
    anchor = 3
    num = 1 
    config = openfile()
    # print(open)    
    vserver = vserver_parse(config)
    service = service_parse(config)
    bind = bind_parse(config)
    servers = server_parse(config)
    for i in vserver: 
        # print(num)
        # print(i)
        # print(anchor)   
        vinfo=i.split(' ')
        num = vserver_create(vinfo, anchor, num)
        v = vinfo[3]
        print(v)
        service_name = find_service(v, bind)
        anchor_return = service_info(service_name, service, servers, anchor)
        merge = cell_merge(anchor, anchor_return)
        anchor = anchor_return
    f_anchor = not_v_service(service, servers, anchor)
    service_stat(f_anchor)
    del anchor
    del f_anchor

        # print(anchor)
        # breakpoint()
        
             
if __name__ == '__main__':
    main()
    print('Done!!!')
    print('Auto SLB to EXCEL')
    print('Create By ShinJH 22.08.24')