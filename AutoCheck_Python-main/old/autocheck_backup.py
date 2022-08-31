import os
import re
import datetime
import sys
# from openpyxl import Workbook
import openpyxl as xl
from model.parse import *
from model.parse2 import *
from model.getrawdata import *
from model.juniper_getdata import *
from model.citrix_getdata import *


##  switch.txt 파일 Array 변환 ##

def get_switch_info(file):
    sw = list
    switch = list()
    f = open(file, 'r')
    if f:
        s = f.read()
    else:
        print('File %s not found' % (file))
    f.close()
    if s:
        sw = s.split('\n')
        for i in sw:
            i = ' '.join(i.split())
            if i != '':
                tmp = i.split(' ')
                switch.append(tmp)
    return switch


def gensimplerep(switch, anchor):
    print(switch)
    wb = xl.load_workbook('report.xlsx')
    ws = wb['report']
    ws['D'+str(anchor)] = switch['model']
    ws['G'+str(anchor)] = switch['hostname']
    ws['L'+str(anchor)] = switch['ip']
    ws['Q'+str(anchor)] = switch['cpu util']
    ws['T'+str(anchor)] = switch['mem util']
    ws['V'+str(anchor)] = switch['temperature']
    ws['Y'+str(anchor)] = switch['power']
    ws['AC'+str(anchor)] = switch['fan']
    wb.save('report.xlsx')
    wb.close()


def repmain():
    now = 'Report Date: ' + str(datetime.datetime.now())
    wb = xl.load_workbook(filename='report.xlsx')
    ws = wb['report']
    ws['A2'] = now
    wb.save('report.xlsx')
    wb.close()
    anchor = 4                           # initial position.
    sw = get_switch_info('switch.txt')
    # print(sw)
    for i in sw:
        print(anchor-3)
        protocol = i[4]
        vendor = i[5]
        if vendor == 'cisco':
            data = CiscoIOS(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch       
        elif vendor == 'juniper':
            data = Juniper(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = juniperParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'citrix':
            data = Citrix(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = citrixParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        else:
            print("Not supported vendor!!\n")

if __name__ == '__main__':
    print('+-------------------------------------------------------------+')
    print('| Auto Check Application                                      |')
    print('| version 0.0.1                                               |')
    print('| Scripted by Shin 2022.06.09                                 |')
    print('+-------------------------------------------------------------+')
    repmain()
    print('Done!!!')
