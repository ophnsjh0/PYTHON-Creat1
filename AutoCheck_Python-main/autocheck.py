import os
import re
import datetime
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment
from parsers.cisco_parse import *
from parsers.cisconx_parse import *
from parsers.juniper_parse import *
from parsers.citrix_parse import *
from parsers.dell_parse import *
from models.cisco_getdata import *
from models.juniper_getdata import *
from models.citrix_getdata import *
from models.cisconx_getdata import *
from models.dasan_getdata import *
from models.arista_getdata import *
from models.piolink_getdata import *
from models.dell_getdata import *
from models.hp_getdata import *
# import openpyxl as xl


##  switch.txt 파일 Array 변환 ##

def get_switch_info(check):
    sw = list
    switch = list()
    if check == "1":
        f = open('scourt_ext.txt', 'r')
    elif check == "2":
        f = open('scourt_int.txt', 'r')
    elif check == "3":
        f = open('iros_ext.txt', 'r')
    elif check == "4":
        f = open('iros_int.txt', 'r')
    elif check == "0":
        f = open('moningcheck.txt', 'r')
    else:
        print("잘못 입력 하였습니다. 다시 실행해 주세요.")
        sys.exit()
    if f:
        s = f.read()
    else:
        print('File %s not found' % (file))
    f.close()
    if s:
        sw = s.split('\n')
        for i in sw:
            i = ' '.join(i.split())
            if i != '':
                tmp = i.split(' ')
                switch.append(tmp)
    return switch


def gensimplerep(switch, anchor):
    # print(switch)
    day = datetime.datetime.now()
    wb = load_workbook('C:/test/report_{0}_{1}_{2}_{3}.xlsx'.format(day.year, day.month, day.day, day.hour))
    ws = wb['report']
    # print(anchor)
    ws['A'+str(anchor)] = switch['model']
    ws['B'+str(anchor)] = switch['hostname']
    ws['C'+str(anchor)] = switch['ip']
    ws['D'+str(anchor)] = switch['cpu util']
    ws['E'+str(anchor)] = switch['mem util']
    ws['F'+str(anchor)] = switch['temperature']
    ws['G'+str(anchor)] = switch['power']
    ws['H'+str(anchor)] = switch['fan']
    box = Border(
        left=Side(border_style="thin", color='00000000'),
        right=Side(border_style="thin", color='00000000'),
        top=Side(border_style="thin", color='00000000'),
        bottom=Side(border_style="thin", color='00000000')
    )
    for rows in ws['A'+str(anchor):'H'+str(anchor)]:
        for cell in rows:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = box
    cpu = ws['D'+str(anchor)].value
    new_cpu = cpu.replace('%', "")
    mem = ws['E'+str(anchor)].value
    new_mem = mem.replace('%', "")
    temp = ws['F'+str(anchor)].value
    new_temp = temp.replace('C', "")
    if isinstance(float(new_cpu), float) and float(new_cpu) > 20:
        ws['D'+str(anchor)].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    if isinstance(float(new_mem), float) and float(new_mem) > 30:
        ws['E'+str(anchor)].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    if isinstance(int(new_temp), int) and int(new_temp) > 30:
        ws['F'+str(anchor)].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    ws.freeze_panes = "A4"
    wb.save('C:/test/report_{0}_{1}_{2}_{3}.xlsx'.format(day.year, day.month, day.day, day.hour))
    wb.close()


def repmain():
    now = 'Report Date: ' + str(datetime.datetime.now())
    day = datetime.datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    ws['A2'] = now
    ws.row_dimensions[3].height = 27
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 25
    ws = wb['report']
    box = Border(
        left=Side(border_style="thin", color='00000000'),
        right=Side(border_style="thin", color='00000000'),
        top=Side(border_style="thin", color='00000000'),
        bottom=Side(border_style="thin", color='00000000')
    )
    for rows in ws["A3":"H3"]:
        for cell in rows:
            cell.fill = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = box
    ws['A3'] = "Model"
    ws['B3'] = "HostName"
    ws['C3'] = "IP Address"
    ws['D3'] = "CPU 사용량"
    ws['E3'] = "MEMORY 사용량"
    ws['F3'] = "Temperature"
    ws['G3'] = "Power 상태" 
    ws['H3'] = "FAN 상태" 
    wb.save('C:/test/report_{0}_{1}_{2}_{3}.xlsx'.format(day.year, day.month, day.day, day.hour))
    wb.close()
    anchor = 4                           # initial position.
    print("###############################################################################################################")
    print("##                                          점검 항목을 선택하여 주세요                                      ##")
    print("##                                                                                                           ##")
    print("##                                           0. 오전점검                                                     ##")
    print("##                                           1. 사법대국민 -> 전자소송                                       ##")
    print("##                                           2. 사법 내부 -> 재판사무                                        ##")
    print("##                                           3. 등기대국민 -> 인터넷 등기소                                  ##")
    print("##                                           4. 등기 내부  -> 신등기                                         ##")
    print("##                                                                                                           ##")
    print("###############################################################################################################")
    check = input("점검 항목을 선택하여 주세요(숫자 1~4 입력) : ")
    sw = get_switch_info(check)
    # print(sw)
    for i in sw:
        print(anchor-3)
        protocol = i[4]
        vendor = i[5]
        if vendor == 'cisco':
            data = CiscoIOS(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch       
        elif vendor == 'juniper':
            data = Juniper(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = juniperParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'citrix':
            data = Citrix(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = citrixParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'cisconx':
            data = CiscoNX(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoNXParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'dasan':
            data = Dasan(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'arista':
            data = Arista(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'dell':
            data = Dell(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = DELLParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'hp':
            data = HP(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        elif vendor == 'piolink':
            data = Piolink(i)
            if protocol == 'ssh':
                raw = data.get_ssh()
            elif protocol == 'telnet':
                raw = data.get_telnet()
            elif protocol == 'tacacs':
                raw = data.get_tacacs()
            else:
                print('Not supported!!!\n')
            del data
            switch = CiscoIOSParse(raw)
            gensimplerep(switch.saveresult(), anchor)
            anchor +=1
            del switch
        else:
            print("Not supported vendor!!\n")

if __name__ == '__main__':
    print('+-------------------------------------------------------------+')
    print('| Auto Check Application                                      |')
    print('| version 0.0.1                                               |')
    print('| Scripted by Shin 2022.06.09                                 |')
    print('+-------------------------------------------------------------+')
    repmain()
    print('Done!!!')
